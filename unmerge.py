from openpyxl import load_workbook

# Path to your original Excel file
file_path = 'Ala wai boys.xlsx'
# Specify the path for the new unmerged file
ungrouped_file_path = 'unmerge.xlsx'

# Load the workbook
wb = load_workbook(filename=file_path)

# Iterate through each worksheet in the workbook
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Unmerge cells
    merged_cells_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_cells_ranges:
        ws.unmerge_cells(str(merged_range))

# Save the workbook to a new file
wb.save(filename=ungrouped_file_path)

print(f"Processed file saved to {ungrouped_file_path}")
